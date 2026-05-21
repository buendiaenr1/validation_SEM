#!/usr/bin/env python3
"""
Script para sustituir textos por numeros en los archivos Excel del Wellness Questionnaire.
Basado en la tabla de mapeo de la imagen proporcionada por el usuario.
Tambien limpia celdas que contienen numeros mezclados con texto (ej: "14 años" -> 14).
Tambien sustituye Femenino -> 0, Masculino -> 1.
Tambien sustituye Si -> 1, No -> 0 en la columna de universitario.
Tambien elimina filas que tengan datos faltantes a partir de la columna de edad.
"""

import re
import openpyxl
from copy import copy

# ============================================================
# TABLA DE MAPEO: Texto -> Numero (escalas Likert)
# ============================================================

MAPEO = {
    # --- Fatiga ---
    "Muy fresco": 5,
    "Recuperado": 5,
    "Fresco": 4,
    "Normal": 3,
    "Más cansado de lo normal": 2,
    "Más fatigado de lo normal": 2,
    "Siempre cansado": 1,
    "Muy fatigado": 1,

    # --- Calidad del sueño ---
    "Sentirse bien": 5,
    "Muy relajante": 5,
    "Bien": 4,
    "Dificultad para conciliar el sueño": 3,
    "Inquieto": 2,
    "Sueño inquieto": 2,
    "Insomnio": 1,

    # --- Dolor muscular general ---
    "Sentirse bien": 5,
    "Muy buenas sensaciones": 5,
    "Bien": 4,
    "Buenas sensaciones": 4,
    "Normal": 3,
    "Aumento de dolor muscular": 2,
    "Muy dolorido": 1,

    # --- Nivel de estrés ---
    "Muy relajado": 5,
    "Relajado": 4,
    "Normal": 3,
    "Estresado": 2,
    "Muy estresado": 1,

    # --- Ánimo ---
    "Estado de ánimo muy positivo": 5,
    "Talante muy positivo": 5,
    "En general, de buen humor": 4,
    "Buen humor": 4,
    "Menos interesado en las actividades de lo normal": 3,
    "Menos interesado otras actividades de lo normal": 3,
    "Irritabilidad hacia compañeros de equipo o familiares": 2,
    "Mal genio": 2,
    "Muy molesto, irritable o deprimido": 1,
}

# ============================================================
# MAPEO DE GÉNERO
# ============================================================

MAPEO_GENERO = {
    "Femenino": 0,
    "Masculino": 1,
}

# ============================================================
# MAPEO DE UNIVERSITARIO
# ============================================================

MAPEO_UNIVERSITARIO = {
    "Sí": 1,
    "Si": 1,
    "No": 0,
}

# Columnas G=7, H=8, I=9, J=10, K=11
COLUMNAS_A_REEMPLAZAR = [7, 8, 9, 10, 11]


def quitar_texto_dejar_numero(valor):
    """
    Si la celda contiene numeros y texto (ej: '14 años'),
    extrae solo el numero. Si ya es numero, lo deja igual.
    """
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return valor
    texto = str(valor).strip()
    match = re.search(r'\d+', texto)
    if match:
        return int(match.group())
    return valor


def fila_tiene_datos_faltantes(ws, fila, col_inicio):
    """
    Verifica si una fila tiene alguna celda vacia
    desde la columna de inicio hasta la columna K (11).
    """
    for col in range(col_inicio, 12):
        if ws.cell(row=fila, column=col).value is None:
            return True
    return False


def procesar_archivo(ruta_entrada, ruta_salida, columnas_limpiar=None,
                     columna_genero=None, columna_universitario=None,
                     col_inicio_verificar=3):
    """
    Abre un archivo Excel y realiza:
    1. Reemplaza textos por numeros en columnas G-K
    2. Limpia numeros mezclados con texto
    3. Sustituye Femenino/Masculino por 0/1
    4. Sustituye Si/No por 1/0 en universitario
    5. Elimina filas con datos faltantes a partir de edad

    Parametros:
        ruta_entrada: ruta del archivo Excel original
        ruta_salida: ruta donde guardar el archivo corregido
        columnas_limpiar: lista de columnas donde quitar texto y dejar solo numeros
        columna_genero: numero de columna donde esta el genero
        columna_universitario: numero de columna donde esta la pregunta de universitario
        col_inicio_verificar: columna desde donde verificar datos faltantes
    """
    print(f"Procesando: {ruta_entrada}")
    wb = openpyxl.load_workbook(ruta_entrada)
    ws = wb[wb.sheetnames[0]]

    reemplazos_ok = 0
    limpiezas_ok = 0
    generos_ok = 0
    universitario_ok = 0
    no_encontrados = {}
    filas_a_eliminar = []

    for fila in range(2, ws.max_row + 1):

        # --- Verificar si la fila tiene datos faltantes ---
        if fila_tiene_datos_faltantes(ws, fila, col_inicio_verificar):
            filas_a_eliminar.append(fila)
            continue  # No procesar esta fila

        # --- Paso 1: Reemplazar textos por numeros en columnas G-K ---
        for col in COLUMNAS_A_REEMPLAZAR:
            celda = ws.cell(row=fila, column=col)
            valor = celda.value
            if valor is None:
                continue
            texto = str(valor).strip()
            if texto in MAPEO:
                celda.value = MAPEO[texto]
                reemplazos_ok += 1
            elif isinstance(valor, str) and texto != "":
                if texto not in no_encontrados:
                    no_encontrados[texto] = []
                no_encontrados[texto].append(celda.coordinate)

        # --- Paso 2: Limpiar numeros mezclados con texto ---
        if columnas_limpiar:
            for col in columnas_limpiar:
                celda = ws.cell(row=fila, column=col)
                valor_original = celda.value
                valor_limpio = quitar_texto_dejar_numero(valor_original)
                if valor_original != valor_limpio and valor_limpio is not None:
                    celda.value = valor_limpio
                    limpiezas_ok += 1

        # --- Paso 3: Sustituir genero Femenino->0, Masculino->1 ---
        if columna_genero:
            celda = ws.cell(row=fila, column=columna_genero)
            valor = celda.value
            if valor is not None:
                texto = str(valor).strip()
                if texto in MAPEO_GENERO:
                    celda.value = MAPEO_GENERO[texto]
                    generos_ok += 1

        # --- Paso 4: Sustituir universitario Si->1, No->0 ---
        if columna_universitario:
            celda = ws.cell(row=fila, column=columna_universitario)
            valor = celda.value
            if valor is not None:
                texto = str(valor).strip()
                if texto in MAPEO_UNIVERSITARIO:
                    celda.value = MAPEO_UNIVERSITARIO[texto]
                    universitario_ok += 1

    # --- Paso 5: Eliminar filas con datos faltantes (de abajo hacia arriba) ---
    for fila in reversed(filas_a_eliminar):
        ws.delete_rows(fila)

    wb.save(ruta_salida)
    wb.close()

    print(f"  Reemplazos texto->numero: {reemplazos_ok}")
    print(f"  Limpiezas de numeros con texto: {limpiezas_ok}")
    print(f"  Reemplazos de genero (F=0, M=1): {generos_ok}")
    print(f"  Reemplazos universitario (No=0, Si=1): {universitario_ok}")
    print(f"  Filas eliminadas por datos faltantes: {len(filas_a_eliminar)}")
    if no_encontrados:
        print(f"  Textos NO encontrados ({len(no_encontrados)}):")
        for texto, celdas in sorted(no_encontrados.items()):
            print(f"    - '{texto}' en {len(celdas)} celda(s)")
    else:
        print("  Todos los textos fueron reemplazados correctamente.")
    print(f"  Archivo guardado: {ruta_salida}")
    return reemplazos_ok, limpiezas_ok, generos_ok, universitario_ok, len(filas_a_eliminar), no_encontrados


def main():
    # Cada archivo tiene las columnas en diferente posicion:
    # - WQT-2:    col C(3)=Edad, col D(4)=Genero, col E(5)=Universitario
    # - WELLNESS: col D(4)=Edad, col E(5)=Genero, col F(6)=Universitario
    archivos = [
        {
            "entrada": "WQT-2 (respuestas).xlsx",
            "salida": "WQT-2 (respuestas)_corregido.xlsx",
            "columnas_limpiar": [3],        # Columna C = Edad
            "columna_genero": 4,            # Columna D = Genero
            "columna_universitario": 5,      # Columna E = Universitario
            "col_inicio_verificar": 3,       # Verificar desde columna C (Edad)
        },
        {
            "entrada": "WELLNESS QUESTIONARY (respuestas) - copia.xlsx",
            "salida": "WELLNESS QUESTIONARY (respuestas) - copia_corregido.xlsx",
            "columnas_limpiar": [4],        # Columna D = Edad
            "columna_genero": 5,            # Columna E = Genero
            "columna_universitario": 6,      # Columna F = Universitario
            "col_inicio_verificar": 4,       # Verificar desde columna D (Edad)
        },
    ]

    for arch in archivos:
        procesar_archivo(
            arch["entrada"],
            arch["salida"],
            arch.get("columnas_limpiar"),
            arch.get("columna_genero"),
            arch.get("columna_universitario"),
            arch.get("col_inicio_verificar")
        )
        print()


if __name__ == "__main__":
    main()
